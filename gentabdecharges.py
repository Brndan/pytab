#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Ce script produit automatiquement les tableaux de décharge
# à partir d’un modèle et d’un fichier composé ainsi :
# | Syndicat | Quotité proposée |
# | Ain (01) | 0,567            |
# | Total    | 2,558            |

# Tous les fichiers, en entrée comme en sortie, sont des XLSX.


import sys
import os, shutil
import argparse

from openpyxl import *


# Fonction qui prend en entrée le fichier donnant les 
# quotités pour tous les syndicats et retourne un
# dictionnaire

def load_syndicats(fichier_decharges):
    try:
        quotite_fichier = load_workbook(filename=fichier_decharges)
    except:
        print("Erreur à l’ouverture du fichier quotité")
        os._exit(1)
        
    ## importer les quotités de décharge dans un dictionnaire
    base_syndicats = {}
    syndicat = quotite_fichier.active

    for ligne in syndicat.iter_rows(min_row=2,
                                    max_row=syndicat.max_row - 1,
                                    min_col=1,
                                    max_col=2,
                                    values_only=True):
        nom_syndicat = ligne[0]
        decharge = ligne[1]
        base_syndicats[nom_syndicat] = decharge
    return base_syndicats


# Fonction qui prend en argument 4 paramètre :
# template_fichier → fichier xlsx modèle ouvert 
# dossier_export → dossier dans lequel les fichiers seront exportés
# nomSyndicat → le nom de chaque structure issu du dictionnaire
# quotiteSyndicat → la quotité de décharge attribuée au syndicat
# 
# À partir de ça : on modifie les deux cellules nécessaires et
# enregistrement en xlsx avec le bon nom de fichier.

def produit_tableau(template_fichier, dossier_export, nomSyndicat,
                    quotiteSyndicat):
    fichier_export = nomSyndicat + ".xlsx"
    template = template_fichier.active
    template["A74"] = nomSyndicat
    template["B74"] = quotiteSyndicat
    template.protection.sheet = True
    template.protection.password = 'pandace'
    template.protection.enable()
    template_fichier.save(
        filename=os.path.join(dossier_export, fichier_export))


def main():
    parser = argparse.ArgumentParser(description="Produire à partir d’un modèle les fichiers de décharge pour tous les syndicat.")
    parser.add_argument(
        "-q",
        "--quotite",
        action="store",
        help=
        "Fichier comprenant la quotité attribuée à chaque syndicat, au format .xlsx",
        required=True)
    parser.add_argument("template",
                        action="store",
                        help="Chemin du fichier modèle au format .xlsx")
    args = parser.parse_args()

    try:
        template_fichier = load_workbook(filename=args.template)
        print("Fichier template ouvert")
    except:
        print("Erreur à l’ouverture du fichier template")
        os._exit(1)

    export = os.path.join(os.getcwd(), 'export')
    if os.path.exists(export):
        try:
            shutil.rmtree(export)
            print("Dossier d’export supprimé")
        except:
            print("Problème lors de la suppression du dossier d’export")
            os_exit(1)
    try:
        os.mkdir(export)
        print("Dossier d’export créé")
    except:
        print("Impossible de créer le dossier d’export")
        os_exit(1)

    base_syndicats = load_syndicats(args.quotite)

    # Pour chaque syndicat, on appelle produit_tableau() qui se charge de 
    # générer le fichier
    for key in base_syndicats:
        produit_tableau(template_fichier, export, key, base_syndicats[key])


if __name__ == "__main__":
    main()
