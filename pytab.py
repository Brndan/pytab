#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Ce script produit automatiquement les tableaux de décharge
# à partir d’un modèle et d’un fichier composé ainsi :
# | Syndicat | Quotité proposée |
# | Ain (01) | 0,567            |
# | Total    | 2,558            |

# Tous les fichiers, en entrée comme en sortie, sont des XLSX.


import sys
import os
import shutil
import argparse

import openpyxl as xlsx


# Fonction qui prend en entrée le fichier donnant les
# quotités pour tous les syndicats et retourne un
# dictionnaire

def load_syndicats(fichier_decharges):
    try:
        quotite_fichier = xlsx.load_workbook(filename=fichier_decharges)
    except:
        sys.exit("Erreur à l’ouverture du fichier quotité")

    # importer les quotités de décharge dans un dictionnaire
    base_syndicats = {}
    syndicat = quotite_fichier.active

    for ligne in syndicat.iter_rows(min_row=2,
                                    max_row=syndicat.max_row - 1,
                                    min_col=1,
                                    max_col=2,
                                    values_only=True):
        nom_syndicat = ligne[0]
        decharge = ligne[1]
        base_syndicats[nom_syndicat] = decharge
    return base_syndicats


# Fonction qui prend en argument 5 paramètres :
# template_fichier → fichier xlsx modèle ouvert
# dossier_export → dossier dans lequel les fichiers seront exportés
# nom_syndicat → le nom de chaque structure issu du dictionnaire
# quotite_syndicat → la quotité de décharge attribuée au syndicat
# sheet_password → l’éventuel mot de passe de la feuille
#
# À partir de ça : on modifie les deux cellules nécessaires et
# enregistrement en xlsx avec le bon nom de fichier.

def produit_tableau(template_fichier, dossier_export, nom_syndicat,
                    quotite_syndicat, sheet_password):
    fichier_export = nom_syndicat + ".xlsx"
    template = template_fichier.active
    template["A64"] = nom_syndicat
    template["B64"] = quotite_syndicat
    if sheet_password:
        template.protection.password = sheet_password
        template.protection.enable()
        template.protection.sheet = True
    template_fichier.save(
        filename=os.path.join(dossier_export, fichier_export))


def main():
    parser = argparse.ArgumentParser(
        description="Produire à partir d’un modèle les fichiers de décharge pour tous les syndicat.")
    parser.add_argument(
        "-q",
        "--quotite",
        action="store",
        help="Fichier comprenant la quotité attribuée à chaque syndicat, au format .xlsx",
        required=True)
    parser.add_argument(
        "-p",
        "--password",
        action="store",
        help="Mot de passe pour protéger les tableaux générés",
        required=False
    )
    parser.add_argument("template",
                        action="store",
                        help="Chemin du fichier modèle au format .xlsx")
    args = parser.parse_args()

    try:
        template_fichier = xlsx.load_workbook(filename=args.template)
        print("Fichier template ouvert", file=sys.stderr)
    except:
        sys.exit("Erreur à l’ouverture du fichier template")

    export = os.path.join(os.getcwd(), 'export')
    if os.path.exists(export):
        try:
            shutil.rmtree(export)
            print("Dossier d’export supprimé", file=sys.stderr)
        except:
            sys.exit("Problème lors de la suppression du dossier d’export")
    try:
        os.mkdir(export)
        print("Dossier d’export créé", file=sys.stderr)
    except:
        sys.exit("Impossible de créer le dossier d’export")

    base_syndicats = load_syndicats(args.quotite)

    # Pour chaque syndicat, on appelle produit_tableau() qui se charge de
    # générer le fichier
    for key in base_syndicats:
        produit_tableau(template_fichier, export, key,
                        base_syndicats[key], args.password)


if __name__ == "__main__":
    main()
